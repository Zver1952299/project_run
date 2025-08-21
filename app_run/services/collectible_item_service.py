from app_run.models import CollectibleItem
from app_run.serializers import CollectibleItemSerializer

from openpyxl import load_workbook


class CollectibleItemService:
    @staticmethod
    def import_from_excel(file):
        broken_rows = []
        wb = load_workbook(file)
        ws = wb.active
        headers = ['name', 'uid', 'value', 'latitude', 'longitude', 'picture']
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            serializer = CollectibleItemSerializer(data=row_dict)
            if serializer.is_valid():
                CollectibleItem.objects.create(**row_dict)
            else:
                broken_rows.append(list(row))
        return broken_rows