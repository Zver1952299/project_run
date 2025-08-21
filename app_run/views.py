from django.core.serializers import serialize
from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.conf import settings
from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404
from django.db.models import Count, Q, Avg, Sum, ExpressionWrapper, F, FloatField, Case, When, Value
from django.db.models.functions import Cast
from .models import Run, AthleteInfo, Challenge, Position, CollectibleItem, Subscribe
from .serializers import RunSerializer, UserSerializer, UserForCollectibleItemSerializer, AthleteInfoSerializer, ChallengeSerializer, PositionSerializer, CollectibleItemSerializer, UserForAthleteSerializer, UserForCoachSerializer, SubscribeSerializer
from .services.run_service import RunService, get_user_or_400
from openpyxl import load_workbook
from haversine import haversine, Unit
from collections import defaultdict


@api_view(['GET'])
def company_details(request):
    details = {
        'company_name': settings.COMPANY_NAME,
        'slogan': settings.SLOGAN,
        'contacts': settings.CONTACTS
    }
    return Response(details)


class RunPagination(PageNumberPagination):
    page_size_query_param = 'size'


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.all().select_related('athlete')
    serializer_class = RunSerializer
    pagination_class = RunPagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete']
    ordering_fields = ['created_at']


class UserPagination(PageNumberPagination):
    page_size_query_param = 'size'


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    pagination_class = UserPagination
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['first_name', 'last_name']
    ordering_fields = ['date_joined']

    def get_queryset(self):
        qs = super().get_queryset().filter(is_superuser=False)

        qs = qs.annotate(
            runs_finished=Count(
                'runs',
                filter=Q(runs__status=Run.Status.FINISHED)
            ),
            rating=Avg('subscribers__rating')
        )

        user_type = self.request.query_params.get('type', None)
        if user_type == 'athlete':
            qs = qs.filter(is_staff=False)
        elif user_type == 'coach':
            qs = qs.filter(is_staff=True)
        return qs

    def get_serializer_class(self):
        if self.action == 'list':
            return UserSerializer
        elif self.action == 'retrieve':
            user = self.get_object()
            if user.is_staff:
                return UserForCoachSerializer
            else:
                return UserForAthleteSerializer
        return super().get_serializer_class()



class RunStatusUpdateView(APIView):
    def post(self, request, id, action):
        try:
            run = RunService.update_status(id, action)
        except ValueError:
            return Response({"detail": "Invalid action"}, status=status.HTTP_400_BAD_REQUEST)
        except RuntimeError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        serializer = RunSerializer(run)
        return Response(serializer.data, status=status.HTTP_200_OK)


class AthleteInfoView(APIView):
    def get_user_or_404(self, id):
        return get_object_or_404(User, id=id)

    def validate_weight(self, weight):
        if weight is None:
            return True
        try:
            weight = int(weight)
        except (TypeError, ValueError):
            return False
        return 0 < weight < 900

    def get_or_create_athlete_info(self, user_id):
        self.get_user_or_404(user_id)

        obj, created = AthleteInfo.objects.get_or_create(
            user_id=user_id,
            defaults={'goals': '', 'weight': None}
        )
        return obj, created

    def build_response(self, obj, created, is_update=False):
        serializer = AthleteInfoSerializer(obj)
        status_code = status.HTTP_201_CREATED if created and is_update else status.HTTP_200_OK
        return Response(serializer.data, status=status_code)

    def get(self, request, id):
        obj, created = self.get_or_create_athlete_info(id)
        return self.build_response(obj, created)

    def put(self, request, id):
        self.get_user_or_404(id)

        weight = request.data.get('weight', None)
        if not self.validate_weight(weight):
            return Response({'detail': 'Weight must be > 0 and < 900'}, status=status.HTTP_400_BAD_REQUEST)
        obj, created = AthleteInfo.objects.update_or_create(
            user_id=id,
            defaults={
                'goals': request.data.get('goals', ''),
                'weight': weight
            }
        )
        return self.build_response(obj, created, is_update=True)


class ChallengeView(APIView):
    def get(self, request):
        qs = Challenge.objects.all()
        athlete = request.query_params.get('athlete')
        if athlete:
            qs = qs.filter(athlete=athlete)
        challenges_list = ChallengeSerializer(qs, many=True).data
        return Response(challenges_list)


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['run']

    def perform_create(self, serializer):
        position = serializer.save()
        athlete = User.objects.get(id=position.run.athlete_id)
        items = CollectibleItem.objects.all()
        for i in items:
            distance = haversine((position.latitude, position.longitude), (i.latitude, i.longitude), unit=Unit.METERS)
            if distance < 100:
                athlete.collectible_items.add(i)

        positions_of_run = self.queryset.filter(run=position.run).values()
        count_positions = len(positions_of_run)
        if count_positions > 1:
            position.distance = round(positions_of_run[count_positions - 2]['distance'] + haversine((positions_of_run[count_positions - 2]['latitude'], positions_of_run[count_positions - 2]['longitude']), (position.latitude, position.longitude)), ndigits=2)
            distance_the_last_points = round(haversine((positions_of_run[count_positions - 2]['latitude'], positions_of_run[count_positions - 2]['longitude']), (position.latitude, position.longitude), unit=Unit.METERS), ndigits=2)
            time = (position.date_time - positions_of_run[count_positions - 2]['date_time']).total_seconds()
            position.speed = round(distance_the_last_points / time, ndigits=2)
            position.save()



class CollectibleItemViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer


class UploadFileView(APIView):
    def post(self, request):
        uploaded_file = request.FILES.get('file')
        broken_rows = []

        if uploaded_file:
            wb = load_workbook(uploaded_file)
            ws = wb.active
            headers = ['name', 'uid', 'value', 'latitude', 'longitude', 'picture']

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                serializer = CollectibleItemSerializer(data=row_dict)

                if serializer.is_valid():
                    CollectibleItem.objects.create(**row_dict)
                else:
                    broken_rows.append(list(row))

        return Response(broken_rows)


class SubscribeView(APIView):
    def post(self, request, id):
        coach = get_object_or_404(User, id=id)
        athlete_id = request.data.get('athlete')
        if not (type(athlete_id) == int):
            if not athlete_id.isdigit():
                return Response({'detail': 'athlete id must be integer'}, status=status.HTTP_400_BAD_REQUEST)
        athlete = get_user_or_400(user_id=athlete_id)
        if not athlete:
            return Response({'detail': 'No User matches the given query.'}, status=status.HTTP_400_BAD_REQUEST)

        if not athlete.is_staff and coach.is_staff:
            subscription, created = Subscribe.objects.get_or_create(athlete=athlete, coach=coach)
            if not created:
                return Response({'detail': 'Already subscribed'}, status=status.HTTP_400_BAD_REQUEST)

            return Response({'detail': 'Subscribed successfully'}, status=status.HTTP_200_OK)

        return Response({'detail': 'Only an athlete can subscribe for a coach.'}, status=status.HTTP_400_BAD_REQUEST)


class ChallengeSummaryView(APIView):
    def get(self, request):
        challenges = Challenge.objects.select_related('athlete').all()
        grouped = defaultdict(list)
        for i in challenges:
            athlete_data = {
                    'id': i.athlete_id,
                    'full_name': i.athlete.first_name + ' ' + i.athlete.last_name,
                    'username': i.athlete.username
            }
            grouped[i.full_name].append(athlete_data)
        data = [
            {'name_to_display': name, 'athletes': athletes}
            for name, athletes in grouped.items()
        ]
        return Response(data)


class RatingView(APIView):
    def post(self, request, coach_id):
        athlete_id = request.data.get('athlete')
        rating = request.data.get('rating')
        try:
            athlete = User.objects.get(id=athlete_id)
        except User.DoesNotExist:
            return Response({
                'detail': 'Invalid athlete_id'
            }, status=status.HTTP_400_BAD_REQUEST)
        coach = get_object_or_404(User, id=coach_id)
        list_athlete_ids = coach.subscribers.values_list("athlete_id", flat=True)
        if int(athlete_id) in list_athlete_ids:
            subscribe = Subscribe.objects.filter(coach_id=coach_id, athlete_id=athlete_id).first()
            if subscribe:
                if rating is None or not str(rating).isdigit():
                    return Response({
                        'detail': 'rating is invalid'
                    }, status=status.HTTP_400_BAD_REQUEST)
                subscribe.rating = rating
                serializer = SubscribeSerializer(subscribe, data={'rating': rating}, partial=True)
                if serializer.is_valid():
                    subscribe.save()
                    return Response({
                        'detail': 'Rating has been saved'
                    })
                else:
                    return Response({'rating': serializer.errors['rating'][0]}, status=status.HTTP_400_BAD_REQUEST)

        else:
            return Response({
                'detail': 'Athlete is not subscribed to coach'
            }, status=status.HTTP_400_BAD_REQUEST)


class AnalyticView(APIView):
    def get(self, request, coach_id):
        athlete_ids = Subscribe.objects.filter(coach_id=coach_id).values_list('athlete_id', flat=True)

        longest_run = (
            Run.objects
            .filter(athlete_id__in=athlete_ids, status=Run.Status.FINISHED)
            .order_by('-distance')
            .values('athlete_id', 'distance')
            .first()
        )

        total_run = (
            Run.objects
            .filter(athlete_id__in=athlete_ids, status=Run.Status.FINISHED)
            .values('athlete_id')
            .annotate(total_distance=Sum('distance'))
            .order_by('-total_distance')
            .first()
        )

        speed_avg = (
            Run.objects
            .filter(athlete_id__in=athlete_ids, status=Run.Status.FINISHED)
            .values('athlete_id')
            .annotate(
                total_distance=Sum("distance"),
                total_time=Sum("run_time_seconds"),
            )
            .annotate(
                avg_speed=(F("total_distance") * 1000) / F("total_time")
            )
            # .order_by('-avg_speed')
            # .first()
        )
        print(f'DEBUG {speed_avg}')
        # print(f'DEBUG {speed_avg['avg_speed']}')


        return Response(
            {
                "longest_run_user": longest_run["athlete_id"] if longest_run else None,
                "longest_run_value": longest_run["distance"] if longest_run else None,

                "total_run_user": total_run["athlete_id"] if total_run else None,
                "total_run_value": total_run["total_distance"] if total_run else None,

                # "speed_avg_user": speed_avg["athlete_id"] if speed_avg else None,
                # "speed_avg_value": speed_avg["avg_speed"] if speed_avg else None,
            }
        )